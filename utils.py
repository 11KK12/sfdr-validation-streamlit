#### Imports
from PyPDF2 import PdfReader
import openai
import pdfplumber
import numpy as np
from numpy.linalg import norm
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
import os
import streamlit as st
import io

# Set up open AI
try:
    openai.api_type = st.secrets["OPENAI_API_TYPE"]
    openai.api_key = st.secrets["OPENAI_API_KEY"]
    openai.api_base = st.secrets["OPENAI_API_BASE"]
    openai.api_version = st.secrets["OPENAI_API_VERSION"]
except:
    openai.api_type = os.environ["OPENAI_API_TYPE"]
    openai.api_key = os.environ["OPENAI_API_KEY"]
    openai.api_base = os.environ["OPENAI_API_BASE"]
    openai.api_version = os.environ["OPENAI_API_VERSION"]

# Functions
def generate_embeddings(text: str) -> list:
    response = openai.Embedding.create(
        input=text, engine="text-embedding-ada-002")
    embeddings = response['data'][0]['embedding']
    return embeddings

def estimate_costs(template_count: int) -> float:
    avg_validation_cost_per_template = 0.07
    custom_extraction_cost_per_page = 0.046817 # with this approach, custom extraction is only used on one page per document -> cost savings
    embedding_costs_per_template = 0.001

    total_costs = (avg_validation_cost_per_template + custom_extraction_cost_per_page + embedding_costs_per_template) * template_count
    
    return total_costs

def get_value(index, name, dataframe):
    try:
        return str(dataframe.at[index, name])
    except:
        return ""

def find_templates_in_pdf(uploaded_file) -> list:
    try:
        pdf_reader = PdfReader(uploaded_file)
        pages = pdf_reader.pages
        num_pages = len(pages)
        templates = []
        i = 0
        last_start_page = 0
        
        # Define all variables that we want to extract (from the first page)
        f_template_article = None
        f_product_name = None
        f_legal_entity_identifier = None

        while i < num_pages:
            page = pages[i]
            text = page.extract_text()
            
            if "asetuksen (eu) 2019/2088" in text.lower():
                # if there has been a previous start page, a new start page means the end of the previous template
                # -> save and reset extracted variables
                if last_start_page != 0:
                    template = {
                        "start_page": last_start_page,
                        "end_page": i,
                        "f_template_article": f_template_article,
                        "f_product_name": f_product_name,
                        "f_legal_entity_identifier": f_legal_entity_identifier
                    }
                    templates.append(template)

                    # Reset variables
                    f_template_article = None
                    f_product_name = None
                    f_legal_entity_identifier = None
            
                # Split text of the first page of the template to find f_template_article, product name, legal identity code
                text = page.extract_text()
                try:
                    f_template_article = int(text.split("2088")[1].split("artiklan")[0].strip())
                except:
                    f_template_article = None
                try:
                    f_product_name = text.split("Tuotenimi")[1].split("Oikeushenkilö")[0].replace(":","").strip()
                except:
                    f_product_name = None
                try:
                    # (legal entity identifier is always 20 digits long; if it is longer, something went wrong in the extraction)
                    f_legal_entity_identifier = text.split("tunnus")[1].split("Ympäristöön")[0].replace(":","").strip()[:20]
                except:
                    # identifier needed for saving and validation
                    if f_product_name != None:
                        f_legal_entity_identifier = f_product_name
                    else:
                        f_legal_entity_identifier = "no_name_found_" + str(i)

                last_start_page = i+1
            i += 1
                
        # add information of last template in PDF to list
        if last_start_page != 0:
            template = {
                "start_page": last_start_page,
                "end_page": num_pages,
                "f_template_article": f_template_article,
                "f_product_name": f_product_name,
                "f_legal_entity_identifier": f_legal_entity_identifier
            }
            templates.append(template) 

        return templates
    
    except:
        return []

def generate_question_embeddings():
    # Define variables for questions (preparation for labelling)
    question_variables = {
        "a_promoted_e_s_characteristics": "Mitä ympäristöön ja/tai yhteiskuntaan littyviä ominaisuuksia tämä rahoitustuote edistää?",
        "a_sustainability_indicators_used": "Mitä kestävyysindikaattoreita käytetään mittaamaan kunkun tämän rahoitustuotteen edistämän ympäristöön tai yhteiskuntaan littyvän ominaisuuden toteutumista?",
        "a_sustainable_investment_objectives": "Mitkä ovat niiden kestävien sijoitusten tavoitteet, jotka rahoitustuotteessa aiotaan tehdä osittain, ja miten kestävä sijoitus edistää näiden tavoitteiden saavuttamista?",
        "a_no_significant_harm": "Miten kestävät sijoitukset, jotka rahoitustuotteessa aiotaan tehdä osittain eivät aiheuta haittaa yhdellekään yynmpäristöön tai yhteiskuntaan liittyvälle kestävälle sijoitustavoitteelle?",          
        "a_principal_adverse_impacts_explaination": "Otetaanko tässä rahoitustuotteessa huomioon pääasialliset haitalliset vaikutukset kestävyystekijöihin?",
        "a_investment_strategy": "Mitä sijoitusstrategiaa tässä rahoitustuotteessa noudatetaan?",
        "a_binding_elements_investment_strategy": "Mitä ovat sijoitusstrategian sitovat osatekijät, joita käytetään valittaessa sijoitukset kunkin tämän rahoitustuotteen edistämän ympäristöön tai yhteiskuntaann littyvän ominaisuuden toteutumiseksi?",
        "a_committed_minimum_rate": "Mikä on sitova vähimmäismäärä, jolla vähennetään niiden sijoitusten laajuuutta, jotka on otettu huomioon ennen sijoitusstrategian soveltamista?",
        "a_policy_good_governance_practice": "Mitkä ovat toimintaperiaatteet, joiden mukaisesti arvioidaan sijoituskohteina olevien yritysten hyviä hallintotapooja?",
        "a_planned_asset_allocation": "Mikä on tälle rahoitustuotteelle suunniteltu varojen allokointi?",
        "a_derivatives_e_s_characteristics": "Miten johdannaisten käyttö saa aikaan rahoitustuotteen edistämien ympäristöön tai yhteiskuntaan liittyvien ominaisuuksien totetumista?",
        "a_minimum_extent_taxonomy_alignment": "Missä määrin kestävät sijoitukset, joillla on ympäristötavoite, ovat EU:n luokitusjärjestelmän mukaisia?", # TODO: form extraction needed here for %?
        "a_invest_fossil_nuclear": "Sijoitetaanko rahoitustuotteessa EU:n luokitusjärjestelmän mukaisiin fossiiliseen kaasuun ja/tai ydinenergiaan liittyviin toimintoihin?", # Tform extraction may be needed here for pie chart? no condition regarding that information though
        "a_minimum_share_env_objective": "Mikä on sellaisten ympäristötavoitteita edistävien kestävien sijoitusten vähimmäisosuus, jotka eivät ole EU:n luokitusjärjestelmän mukaisia?",
        "a_minimum_share_social_investment": "Mikä on yhteiskunnallisesti kestävien sijoitusten vähimmäisosuus?",
        "a_investment_included_in_other": "Mitkä sijoitukset sisältyvät kohtaan “#2 Muu”, mikä on niiden tarkoitus ja sovelletaanko ympäristöön liittyvi tai yhteiskunnallisia vähimmäistason suojatoimia?",
        "a_specific_index_benchmark": "Onko tietty indeksi nimetty vertailuarvoksi, jotta voidaanmäärittää, vastaako tämä rahoitustuote edistämiään ympäristöön ja/tai yhteiskuntaan liittyviä ominaisuuksia?",
        "a_product_information_online": "Mistä voin saada tarkempia tuotekohtaisia tietoja verkossa?"
    }
    # generate embedding for each question variable
    for key, value in question_variables.items():
        emb = generate_embeddings(value)
        question_variables[key] = emb
    
    return question_variables

def extract_template_data(template, uploaded_file, document_analysis_client, question_variables):
    # Get list of all paragraphs in a template with style information
    paragraph_list = []
    pages_text = {}
    fontname_dict = {}
    size_dict = {}

    with pdfplumber.open(uploaded_file) as pdf: 
        current_paragraph = ""
        paragraph_fontname = None
        paragraph_size = None
        page_number = None

        # Consider all pages from template
        # PDF page numbers start at 1 but pdf.pages starts with index 0 --> start_page -1
        for i in range(template["start_page"]-1, template["end_page"]):

            page = pdf.pages[i]
            page_text = ""

            for char in page.chars:
                page_text += char["text"]   
                fontname = char["fontname"]
                fontname_dict[fontname] = fontname_dict.get(fontname, 0) + 1
                size = round(char["size"])
                size_dict[size] = size_dict.get(size, 0) + 1

                if (fontname == paragraph_fontname) & (size == paragraph_size):
                    current_paragraph += char["text"]
                else:
                    if current_paragraph != "":
                        paragraph_list.append({"text": current_paragraph.strip(), "fontname": paragraph_fontname, "size": paragraph_size, "page": page_number})
                    current_paragraph = char["text"]
                    paragraph_fontname = fontname
                    paragraph_size = size
                    page_number = char["page_number"]

            if current_paragraph != "":
                paragraph_list.append({"text": current_paragraph.strip(), "fontname": paragraph_fontname, "size": paragraph_size, "page": page_number})

            pages_text[page.page_number] = page_text

    # Select all paragraphs that are bold and contain a questionmark
    question_list = []
    main_fontname = max(fontname_dict, key = fontname_dict.get)
    main_size = max(size_dict, key = size_dict.get)

    for paragraph in paragraph_list:
        # not all questions are bold -> see Alandsbanken; however, not all paragraphs that contain question marks are actual template questions either -> see Nordea
        # --> relevant paragraph if: contains "?" AND (font is bold OR size is bigger than main size OR font is other than main font)
        if ("?" in paragraph["text"]) and (("bold" in paragraph["fontname"].lower()) or (paragraph["size"] != main_size) or (paragraph["fontname"] != main_fontname)):
            # Problem: question on top of the template box ("Onko tällä rahoitustuotteella kestävä sijoitustavoite?") not extracted in proper order
            if not "onko tällä rahoitustuotteella kestävä sijoitustavoite" in paragraph["text"].lower():
                question_list.append(paragraph)

    # Get all text between two questions as associated answers
    q_n_a_pairs = {}

    for i in range(len(question_list)-1):

        # get text split between current question and next question
        start_question = question_list[i]["text"]
        stop_question = question_list[i+1]["text"]

        # only consider relevant pages of the template to reduce risk of wrong splitting (sometimes questions are refered to in answer to other question)
        first_page = question_list[i]["page"]
        last_page = question_list[i+1]["page"]

        relevant_text = " " # relevant text starts with white space to ensure that the start of the relevant text is never equal to the start question, so that the the answer will always be in split[1], not split[0]
        for j in range(first_page, last_page+1):
            relevant_text += pages_text[j] + " "

        answer = relevant_text.split(start_question)[1].split(stop_question)[0].strip()

        q_n_a_pairs[start_question] = answer

    # get answer for last question
    last_question = question_list[len(question_list)-1]["text"]
    first_page = question_list[len(question_list)-1]["page"]
    last_page = template["end_page"]

    relevant_text = " "
    for j in range(first_page, last_page+1):
        relevant_text += pages_text[j] + " "

    answer = relevant_text.split(last_question)[1].strip() # last answer is everything from last question to end of template

    q_n_a_pairs[last_question] = answer

    # Match questions (and answers) with variables of interested for template validation using embeddddings
    # generate and match embedding for each extracted question with question variable embeddings
    for question, answer in q_n_a_pairs.items():

        quest_emb = generate_embeddings(question)

        for key, value in question_variables.items():

            # calculate cosine similarity between value and quest emb
            cosine = np.dot(quest_emb,value)/(norm(quest_emb)*norm(value))
            
            # if variable above treshhold, add to template dict with cosine similarity; if already in dict, keep value with higher cosine similarity
            #if cosine > 0.85: # TODO: no treshhold here? get a label for each extracted q_n_a pair and then check later which can be used
            if cosine > 0:
                # (this way it is possible that multiple questions have the same label --> keep in mind for validation)
                current_value = template.get(question, {})
                current_cosine = current_value.get("cosine",0)

                if cosine > current_cosine:
                    template[question] = {"label": key, "answer": answer, "cosine": cosine}
    
    # Use Azure AI Document Intelligence to get labeled fields from table on first page ####################
    start_page = template["start_page"]

    #with open(uploaded_file, "rb") as f:
    poller = document_analysis_client.begin_analyze_document(
        "sfdr_template_extraction_paid_version_only_1_page",
        document=uploaded_file.getvalue(),
        pages=start_page
    )
    result = poller.result()

    for document in result.documents:
        for k, v in document.fields.items():
            template[k] = v.value

    return template

def template_checks_to_excel(tempys, template_checks):
    #Export validation results into structured excel file
    val_result_dfs = {}

    for k,v in template_checks.items():
        # write results to df
        val_results = pd.DataFrame(None, [], ["name", "description", "value", "comment"])

        for i in range(len(v)):
            val_results.at[i, "name"] = v[i]["name"]
            val_results.at[i, "description"] = v[i]["description"]
            val_results.at[i, "value"] = v[i]["value"]
            val_results.at[i, "comment"] = v[i]["comment"]

        val_result_dfs[k] = val_results

    # buffer to use for excel writer
    buffer = io.BytesIO()

    #with pd.ExcelWriter(filename) as writer:
    with pd.ExcelWriter(buffer) as writer:
        for k, df in val_result_dfs.items():
            extracted_data = tempys[tempys["f_legal_entity_identifier"] == k].transpose()
            if len(k) > 30: # excel sheet name cant be longer than 30 characters
                k = k[len(k)-30:]
            df.to_excel(writer, sheet_name=k, index=False)
            extracted_data.to_excel(writer, sheet_name=k, header=[k], startrow=15)
            
    return buffer

def change_excel_design(buffer):
    
    wb = load_workbook(buffer)

    for ws in wb._sheets:
        
        # change column width
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 50

        # merge B (2) to D (4) for template data rows
        for i in range(18,67):
            ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)

        # add conditional formatting for True / False
        green_fill = PatternFill(start_color="92D050",end_color="92D050",fill_type="solid")
        red_fill = PatternFill(start_color="C00000",end_color="C00000",fill_type="solid")

        ws.conditional_formatting.add('C2:C15', CellIsRule(operator='equal', formula=["True"], fill=green_fill))
        ws.conditional_formatting.add('C2:C15', CellIsRule(operator='equal', formula=["False"], fill=red_fill))
        
        # style to add borders
        border_style = Side(border_style="medium", color="000000")

        true_count = 0
        false_count = 0

        for row in ws.iter_rows():
            for cell in row:

                # count true and false values (for tab formatting)
                if cell.value == True:
                    true_count += 1
                elif cell.value == False:
                    false_count += 1

                # change text alignment
                cell.alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
                cell.font = Font(size=11)
                cell.border = None

                # add borders
                if cell.row == 1:
                    cell.font = Font(size=11, bold=True)
                    if cell.column == 4:
                        cell.border = Border(top=border_style, left=None, right=border_style, bottom=border_style)
                    else:
                        cell.border = Border(top=border_style, left=None, right=None, bottom=border_style)
                elif cell.row == 65:
                    if cell.column == 4:
                        cell.border = Border(top=None, left=None, right=border_style, bottom=border_style)
                    else:
                        cell.border = Border(top=None, left=None, right=None, bottom=border_style)
                elif (cell.column == 4) & (cell.row <= 65):
                    cell.border = Border(top=None, left=None, right=border_style, bottom=None)

        # add headers
        # "VALIDATION"
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        cell = ws.cell(row=1, column=1)
        cell.value = "VALIDATION"
        cell.alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
        cell.font = Font(size=16, bold=True)
        cell.border = Border(top=border_style, left=border_style, right=None, bottom=border_style)
        ws.cell(row=1, column=2).border = Border(top=border_style, left=None, right=None, bottom=border_style)
        ws.cell(row=1, column=3).border = Border(top=border_style, left=None, right=None, bottom=border_style)
        ws.cell(row=1, column=4).border = Border(top=border_style, left=None, right=border_style, bottom=border_style)
        ws.row_dimensions[1].height = 40

        # "TEMPLATE DATA"
        ws.delete_rows(17)
        ws.insert_rows(17)
        ws.merge_cells(start_row=17, start_column=1, end_row=17, end_column=4)
        cell = ws.cell(row=17, column=1)
        cell.value = "TEMPLATE DATA"
        cell.alignment = Alignment(wrap_text=True, horizontal="center",vertical="center")
        cell.font = Font(size=16, bold=True)
        cell.border = Border(top=border_style, left=border_style, right=None, bottom=border_style)
        ws.cell(row=17, column=2).border = Border(top=border_style, left=None, right=None, bottom=border_style)
        ws.cell(row=17, column=3).border = Border(top=border_style, left=None, right=None, bottom=border_style)
        ws.cell(row=17, column=4).border = Border(top=border_style, left=None, right=border_style, bottom=border_style)
        ws.row_dimensions[17].height = 40
        
        # Change tab color based on validation results
        if false_count == 0:
            # green if all validations positive
            ws.sheet_properties.tabColor = "92D050"
        elif true_count == 0:
            # red if all validations negative
            ws.sheet_properties.tabColor = "C00000" 
        else:
            ratio = true_count / false_count
            # yellow if at least 50 % positive
            if ratio >= 1:
                ws.sheet_properties.tabColor = "FFC000"
            else:
                # red if less than 50 % positive
                ws.sheet_properties.tabColor = "C00000"
           
    file_buffer = io.BytesIO()
    wb.save(file_buffer)
    content = file_buffer.getvalue()
        
    return content
